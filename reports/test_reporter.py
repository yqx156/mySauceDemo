"""
测试报告生成器
"""
import os
from datetime import datetime
from dataclasses import dataclass
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

from core.logger_config import logger

@dataclass
class TestResult:
    """测试结果数据类"""
    test_name: str
    username: str
    status: str
    execution_time: str
    error_message: str = ""
    description: str = ""

class TestReporter:
    """测试报告生成器"""
    
    def __init__(self):
        self.test_results: List[TestResult] = []
    
    def add_test_result(self, result: TestResult):
        """添加测试结果"""
        self.test_results.append(result)
        logger.debug(f"添加测试结果: {result.test_name} - {result.username} - {result.status}")
    
    def save_results_to_excel(self) -> str:
        """保存测试结果到Excel文件"""
        try:
            # 创建报告目录
            reports_dir = "test_reports"
            if not os.path.exists(reports_dir):
                os.makedirs(reports_dir)
            
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_results_{timestamp}.xlsx"
            filepath = os.path.join(reports_dir, filename)
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 创建详细结果工作表
            self._create_detailed_results_sheet(wb)
            
            # 创建汇总统计工作表
            self._create_summary_sheet(wb)
            
            # 创建按功能分组的工作表
            self._create_function_summary_sheet(wb)
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 保存文件
            wb.save(filepath)
            logger.info(f"Excel测试报告已保存: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"保存Excel报告失败: {str(e)}")
            return ""
    
    def _create_detailed_results_sheet(self, wb):
        """创建详细结果工作表"""
        ws = wb.active
        ws.title = "详细测试结果"
        
        # 设置表头
        headers = ["测试功能", "用户名", "测试状态", "执行时间", "错误信息", "功能描述"]
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加数据行
        for result in self.test_results:
            row_data = [
                self._clean_text(result.test_name),
                self._clean_text(result.username),
                result.status,
                result.execution_time,
                self._clean_text(result.error_message),
                self._clean_text(result.description)
            ]
            ws.append(row_data)
        
        # 设置数据行样式
        for row_num in range(2, len(self.test_results) + 2):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                
                # 根据测试状态设置颜色
                if col_num == 3:  # 状态列
                    if cell.value == "PASSED":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell.value == "FAILED":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                
                # 设置对齐方式
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 自适应列宽
        column_widths = [25, 15, 12, 20, 40, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    def _create_summary_sheet(self, wb):
        """创建汇总统计工作表"""
        ws = wb.create_sheet("汇总统计")
        
        # 计算统计数据
        summary = self.get_test_summary()
        
        # 设置标题
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "测试执行汇总统计"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加统计信息
        stats = [
            ["总测试数", summary["total"]],
            ["通过测试数", summary["passed"]],
            ["失败测试数", summary["failed"]],
            ["通过率", f"{summary['pass_rate']:.2f}%"],
            ["", ""],  # 空行
            ["执行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        row_start = 3
        for i, (label, value) in enumerate(stats, row_start):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            
            # 设置样式
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].alignment = Alignment(horizontal="right")
            ws[f'B{i}'].alignment = Alignment(horizontal="left")
            
            # 通过率特殊着色
            if label == "通过率":
                if summary['pass_rate'] >= 90:
                    ws[f'B{i}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif summary['pass_rate'] >= 70:
                    ws[f'B{i}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws[f'B{i}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 按用户统计
        user_stats = self._get_user_statistics()
        if user_stats:
            ws[f'A{row_start + len(stats) + 1}'] = "按用户统计："
            ws[f'A{row_start + len(stats) + 1}'].font = Font(bold=True, size=14)
            
            user_row_start = row_start + len(stats) + 3
            ws[f'A{user_row_start}'] = "用户名"
            ws[f'B{user_row_start}'] = "总测试"
            ws[f'C{user_row_start}'] = "通过"
            ws[f'D{user_row_start}'] = "失败"
            ws[f'E{user_row_start}'] = "通过率"
            
            # 设置表头样式
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{user_row_start}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            for i, (username, stats) in enumerate(user_stats.items(), user_row_start + 1):
                ws[f'A{i}'] = username
                ws[f'B{i}'] = stats['total']
                ws[f'C{i}'] = stats['passed']
                ws[f'D{i}'] = stats['failed']
                ws[f'E{i}'] = f"{stats['pass_rate']:.1f}%"
        
        # 设置列宽
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _create_function_summary_sheet(self, wb):
        """创建按功能分组的工作表"""
        ws = wb.create_sheet("功能测试统计")
        
        # 按功能分组统计
        function_stats = self._get_function_statistics()
        
        # 设置标题
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "按测试功能统计"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置表头
        headers = ["测试功能", "总执行次数", "通过次数", "失败次数", "通过率"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加功能统计数据
        row_num = 4
        for func_name, stats in function_stats.items():
            ws.cell(row=row_num, column=1, value=func_name)
            ws.cell(row=row_num, column=2, value=stats['total'])
            ws.cell(row=row_num, column=3, value=stats['passed'])
            ws.cell(row=row_num, column=4, value=stats['failed'])
            ws.cell(row=row_num, column=5, value=f"{stats['pass_rate']:.1f}%")
            
            # 根据通过率设置颜色
            pass_rate_cell = ws.cell(row=row_num, column=5)
            if stats['pass_rate'] == 100:
                pass_rate_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif stats['pass_rate'] >= 50:
                pass_rate_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                pass_rate_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row_num += 1
        
        # 设置列宽
        column_widths = [30, 15, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    def _get_user_statistics(self) -> Dict:
        """获取按用户的统计信息"""
        user_stats = {}
        
        for result in self.test_results:
            username = result.username
            if username not in user_stats:
                user_stats[username] = {'total': 0, 'passed': 0, 'failed': 0}
            
            user_stats[username]['total'] += 1
            if result.status == "PASSED":
                user_stats[username]['passed'] += 1
            else:
                user_stats[username]['failed'] += 1
        
        # 计算通过率
        for username in user_stats:
            total = user_stats[username]['total']
            passed = user_stats[username]['passed']
            user_stats[username]['pass_rate'] = (passed / total * 100) if total > 0 else 0
        
        return user_stats
    
    def _get_function_statistics(self) -> Dict:
        """获取按功能的统计信息"""
        function_stats = {}
        
        for result in self.test_results:
            func_name = result.test_name
            if func_name not in function_stats:
                function_stats[func_name] = {'total': 0, 'passed': 0, 'failed': 0}
            
            function_stats[func_name]['total'] += 1
            if result.status == "PASSED":
                function_stats[func_name]['passed'] += 1
            else:
                function_stats[func_name]['failed'] += 1
        
        # 计算通过率
        for func_name in function_stats:
            total = function_stats[func_name]['total']
            passed = function_stats[func_name]['passed']
            function_stats[func_name]['pass_rate'] = (passed / total * 100) if total > 0 else 0
        
        return function_stats
    
    def _clean_text(self, text: str) -> str:
        """清理文本，移除不合适的字符"""
        if not text:
            return ""
        
        # 移除ANSI转义序列
        ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
        cleaned = ansi_escape.sub('', str(text))
        
        # 移除Excel不支持的控制字符
        cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', cleaned)
        
        # 限制长度
        if len(cleaned) > 300:
            cleaned = cleaned[:300] + "...(已截断)"
        
        return cleaned
    
    def get_test_summary(self) -> Dict:
        """获取测试摘要统计"""
        total = len(self.test_results)
        if total == 0:
            return {"total": 0, "passed": 0, "failed": 0, "pass_rate": 0.0}
        
        passed = sum(1 for result in self.test_results if result.status == "PASSED")
        failed = total - passed
        pass_rate = (passed / total) * 100
        
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": pass_rate
        }
    
    def clear_results(self):
        """清空测试结果"""
        self.test_results.clear()
        logger.info("测试结果已清空")

# 全局测试报告实例
test_reporter = TestReporter()